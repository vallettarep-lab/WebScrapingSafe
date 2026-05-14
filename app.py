import streamlit as st
import pandas as pd
import re
import requests
import time
from io import BytesIO
from openpyxl import load_workbook
from openpyxl.styles import Alignment,Font

headers = {
    "User-Agent": "Mozilla/5.0"
}

def url_exists(url):
    try:
        response = requests.head(
            url,
            headers=headers,
            timeout=5,
            allow_redirects=True
        )

        return response.status_code == 200

    except requests.RequestException:
        return False
    
def normalize_text(text,remove_words):
    # 小文字化
    text = text.lower()

    # アポストロフィ削除
    text = text.replace("'", "").replace("’", "")

    # 英数字以外をスペースに置換
    text = re.sub(r'[^a-z0-9]+', ' ', text)

    # 単語分割
    words = text.split()

    # 削除したい単語
    #remove_words = {"a", "the", "to", "from", "in", "of", "and"}

    # フィルタ
    words = [w for w in words if w not in remove_words]

    # ハイフン連結
    result = '-'.join(words)

    return result

def searchDate(lines, index, pattern):
  recordLength = 12
  for i in range(recordLength):
    date = lines[index+i]
    if re.search(pattern, date):
      return i
  return -1

#AtlanticCouncil
def get_atlanticCouncil_articles(text):
    lines = text.splitlines()
    index = -1;
    target = "Content"
    if target in lines:
        index = lines.index(target)
    else:
        st.error("キーワード「Content」が見つかりませんでした。")
        return None
    rows = []

    pattern = r'^[A-Z][a-z]+ \d{1,2}, \d{4}$'
    while True:
        resultSearchDate = searchDate(lines,index, pattern)
        if resultSearchDate == -1:
            break
        else:
            index += resultSearchDate
        date = lines[index]
        tag = lines[index-2]
        if(tag == "In the News"):
            index += 7
            continue
        title = lines[index+1]
        author = lines[index+2]
        if author.startswith("By"):
            author = author[2:].strip()
            overview = lines[index+4]
            topic1 = lines[index+6]
            topic2 = lines[index+7]
        else:
            author = "N/A"
            overview = lines[index+2]
            topic1 = lines[index+4]
            topic2 = lines[index+5]
        fullurl = "N/A"
        remove_words = {""}
        url = normalize_text(title,remove_words)
        if(tag == "Dispatches"):
            fullurl = f"https://www.atlanticcouncil.org/dispatches/{url}"
        elif(tag == "Issue Brief"):
            fullurl = f"https://www.atlanticcouncil.org/in-depth-research-reports/issue-brief/{url}"
        elif(tag == "UkraineAlert"):
            fullurl = f"https://www.atlanticcouncil.org/blogs/ukrainealert/{url}"
        elif(tag == "MENASource"):
            fullurl = f"https://www.atlanticcouncil.org/blogs/menasource/{url}"
        elif(tag == "Transcript"):
            fullurl = f"https://www.atlanticcouncil.org/news/transcripts/{url}"
        elif(tag == "Event Recap"):
            fullurl = f"https://www.atlanticcouncil.org/commentary/event-recap/{url}"
        else:
            tagurl = normalize_text(tag,remove_words)
            fullurl = f"https://www.atlanticcouncil.org/content-series/{tagurl}/{url}"

        if(not url_exists(fullurl)):
            fullurl = "N/A"

        rows.append(["", date, title, fullurl,"Atlantic Council","",f"{topic1}、{topic2}",author,"",overview])
        index += 7
        time.sleep(0.5)
    df = pd.DataFrame(rows, columns=["#", "日付", "レポートタイトル", "URL","Thinktank名","関係国","トピック","執筆者","まとめ翻訳","まとめ翻訳英文"])
    return df

#CSIS
def get_csis_articles(text):
    lines = text.splitlines()
    index = -1;
    target = "Display Archived Results"
    if target in lines:
        index = lines.index(target)
    else:
        st.error("キーワード「Display Archived Results」が見つかりませんでした。")
        return None

    rows = []
    pattern = r'(?:by\s+)?(.+?)\s+[—-]\s+([A-Z][a-z]+ \d{1,2}, \d{4})'
    while True:
        resultSearchDate = searchDate(lines,index,pattern)
        if resultSearchDate == -1:
            break
        else:
            index += resultSearchDate
        title = lines[index-3]

        match = re.search(r'([A-Z][a-z]+ \d{1,2}, \d{4})', lines[index])

        if match:
            date = match.group(1)
        else:
            date = None

        match = re.search(r'by (.*?) —', lines[index])

        if match:
            author = match.group(1)
        else:
            author = "N/A"

        overview = lines[index-2]

        rows.append(["", date, title, "","CSIS","","",author,"",overview])
        index += 3
    df = pd.DataFrame(rows, columns=["#", "日付", "レポートタイトル", "URL","Thinktank名","関係国","トピック","執筆者","まとめ翻訳","まとめ翻訳英文"])

    return df

#Brookings
def get_brookings_articles(text):
    lines = text.splitlines()
    
    index = -1
    pattern = re.compile(r"\d{1,7}\s+results found")

    for i, line in enumerate(lines):
        if pattern.search(line):
            index = i
            break
    if index == -1:
        st.error("キーワード「数字 results found」が見つかりませんでした。")
        return None
    
    rows = []
    pattern = r'[A-Z][a-z]+ \d{1,2}, \d{4}$'
    while True:
        resultSearchDate = searchDate(lines,index,pattern)
        if resultSearchDate == -1:
            break
        else:
            index += resultSearchDate
        date = lines[index]
        pattern = r'[A-Z][a-z]+ \d{1,2}, \d{4}'
        match = re.search(pattern, date)
        date = match.group()
        title = lines[index-3]
        title2 = lines[index-6]
        if title == title2:
            author = lines[index-2]
            topic = lines[index-4]
        else:
            author = "N/A"
            topic = lines[index-3]
            title = lines[index-2]

        rows.append(["", date, title, "","Brookings Institute","",topic,author,"",""])
        index += 5
    df = pd.DataFrame(rows, columns=["#", "日付", "レポートタイトル", "URL","Thinktank名","関係国","トピック","執筆者","まとめ翻訳","まとめ翻訳英文"])

    return df

#CFR
def get_cfr_articles(text):
    lines = text.splitlines()
    
    index = -1
    pattern = re.compile(r"\d{1,7}\s+results")

    for i, line in enumerate(lines):
        if pattern.search(line):
            index = i
            break
    if index == -1:
        st.error("キーワード「数字 results」が見つかりませんでした。")
        return None
    
    rows = []
    pattern = r'^[A-Z][a-z]+ \d{1,2}, \d{4}$'
    while True:
        resultSearchDate = searchDate(lines,index,pattern)
        if resultSearchDate == -1:
            break
        else:
            index += resultSearchDate
        date = lines[index]
        title = lines[index-1]
        author = lines[index+1]
        if author.startswith("By"):
            author = author[2:].strip()
        else:
            author = "N/A"

        rows.append(["", date, title, "","CFR","","N/A",author,"",""])
        index += 3
    df = pd.DataFrame(rows, columns=["#", "日付", "レポートタイトル", "URL","Thinktank名","関係国","トピック","執筆者","まとめ翻訳","まとめ翻訳英文"])

    return df

#Hudson
def get_hudson_articles(text):
    lines = text.splitlines()
    
    index = -1
    pattern = re.compile(r"\d{1,7}\s+Results:")

    for i, line in enumerate(lines):
        if pattern.search(line):
            index = i
            break
    if index == -1:
        st.error("キーワード「数字 Results:」が見つかりませんでした。")
        return None
    
    rows = []
    pattern = r'^[A-Z][a-z]+ \d{1,2}, \d{4}$'
    while True:
        resultSearchDate = searchDate(lines,index,pattern)
        if resultSearchDate == -1:
            break
        else:
            index += resultSearchDate
        date = lines[index]
        title = lines[index-2]
        author = lines[index-1]
        topic = lines[index-3]

        fullurl = "N/A"
        remove_words = {"a", "the", "to", "from", "in", "of", "and"}
        words = topic.split()
        if(len(words) > 5):
            topic = "N/A"
            url = normalize_text(f"{title} {author}",remove_words)
            fullurl = f"https://www.hudson.org/{url}"
        else:
            url = normalize_text(f"{title} {author}",remove_words)
            urlTopic = normalize_text(topic,remove_words)
            fullurl = f"https://www.hudson.org/{urlTopic}/{url}"

        if(not url_exists(fullurl)):
            fullurl = "N/A"

        rows.append(["", date, title, fullurl,"Hudson","",topic,author,"",""])
        index += 5
        time.sleep(0.5)
    df = pd.DataFrame(rows, columns=["#", "日付", "レポートタイトル", "URL","Thinktank名","関係国","トピック","執筆者","まとめ翻訳","まとめ翻訳英文"])

    return df

# クリア処理を行う関数を定義
def clear_text():
    st.session_state["input_text"] = ""

# ------------------------
# パスワード設定
# ------------------------
PASSWORD = st.secrets["password"] 

if "authenticated" not in st.session_state:
    st.session_state.authenticated = False

# ------------------------
# ログイン画面
# ------------------------
if not st.session_state.authenticated:

    st.title("ログイン")

    password_input = st.text_input("パスワードを入力してください", type="password")

    if st.button("ログイン"):
        if password_input == PASSWORD:
            st.session_state.authenticated = True
            st.rerun()
        else:
            st.error("パスワードが違います")

# ------------------------
# ログイン後の画面
# ------------------------
else:

    st.title("Web記事取得ツール")

    st.write("テキストを入力して Web の記事一覧を Excel で取得できます")

    site = st.selectbox(
        "取得するサイトを選択してください",
        ["Atlantic Council","CSIS","Brookings Institute","CFR","Hudson"]
    )

    # レイアウト作成
    col1, col2 = st.columns([9, 1], vertical_alignment="bottom")

    with col1:
        # テキストエリア。valueではなくkeyで管理します。
        # ※初期値を与えたい場合は、ここではなくsession_stateの初期化で行います。
        text = st.text_area(
            "Webサイトから取得したテキストを入力してください", 
            key="input_text",
            height=200
        )

    with col2:
        # ボタンの on_click に関数を渡す（ここがポイント！）
        st.button("クリア", on_click=clear_text, use_container_width=True)

    order = st.radio(
        "並び順",
        ["新しい順（サイトそのまま）", "古い順（順番反転）"]
    )
    
    # ボタン
    if st.button("記事を取得"):

        st.write("取得中です...")

        df = None

        try:
            if site == "Atlantic Council":
                df = get_atlanticCouncil_articles(text)
            elif site == "CSIS":
                df = get_csis_articles(text)
            elif site == "Brookings Institute":
                df = get_brookings_articles(text)
            elif site == "CFR":
                df = get_cfr_articles(text)
            elif site == "Hudson":
                df = get_hudson_articles(text)

            if df is not None:
                if len(df) > 0:
                    st.success(f"{len(df)}件の記事を取得しました")
                else:
                    st.error("記事がありませんでした")

        except Exception as e:
            st.error(str(e))
        
        if order == "古い順（順番反転）":
            df = df.iloc[::-1].reset_index(drop=True)
        
        if df is not None:
            # Excel作成
            output = BytesIO()
            df.to_excel(output, index=False, engine="openpyxl")

            # openpyxlで加工するために再読み込み
            output.seek(0)
            wb = load_workbook(output)
            ws = wb.active

            # ------------------------
            # 列幅調整
            # ------------------------
            col_widths = {
                "A": 3,
                "B": 12,
                "C": 40,
                "D": 40,
                "E": 28,
                "F": 20,
                "G": 20,
                "H": 20,
                "I": 67,
                "J": 67
            }

            for col, width in col_widths.items():
                ws.column_dimensions[col].width = width

            font = Font(name="Meiryo UI")

            # ------------------------
            # 折り返し設定、中央揃え、フォント設定
            # ------------------------
            for row in ws.iter_rows():
                for cell in row:
                    cell.alignment = Alignment(
                        wrap_text=True,
                        vertical="center"
                    )
                    cell.font = font
            # ------------------------
            # URLにリンク付与
            # ------------------------
            from openpyxl.styles import Font

            for row in ws.iter_rows(min_row=2):
                cell = row[3]  # D列
                if cell.value:
                    cell.hyperlink = cell.value
                    cell.font = Font(name="Meiryo UI", color="0000FF", underline="single")
            # ------------------------
            # 再保存（重要）
            # ------------------------
            output2 = BytesIO()
            wb.save(output2)
            excel_data = output2.getvalue()

            # ダウンロード
            st.download_button(
                label="Excelダウンロード",
                data=excel_data,
                file_name=f"{site}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            )
            